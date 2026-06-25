#!/usr/bin/env python3
"""
Schema validator for normalized AI cost output.

Reads the allowed values from ai-cost-allocation-workbook.xlsx (Enumerations tab)
and checks that every enum-controlled field in a normalized CSV carries a valid
value. Proves output is schema-VALID, not just schema-COMPLETE.

Field handling:
  - HARD fields  -> value MUST be in the canonical enum (else error; --check exits 1)
  - SOFT fields  -> PricingUnit; native billing units vary, so mismatches WARN only
  - Sentinels    -> "" and "n/a" are always allowed (field not applicable to the line)
  - Matching     -> case-insensitive; compound enum entries ("a / b / c") split to atoms

Usage:
    python validate.py <normalized.csv> [--workbook ai-cost-allocation-workbook.xlsx] [--check]
"""
import csv, os, sys, difflib

HARD_FIELDS = [
    "ChargeCategory","DeploymentModel","ModelChannel","ModelTier","CostDomain",
    "WorkloadLifecycle","PricingModel","CommitmentDiscountType","CommitmentDiscountStatus",
    "DirectVsAllocated","FixedVsVariable","CapexVsOpex","Environment","GovernanceStatus","WasteFlag",
]
SOFT_FIELDS = ["PricingUnit"]
SENTINELS = {"", "n/a"}

DEFAULT_WB = os.path.join(os.path.dirname(__file__), "..", "ai-cost-allocation-workbook.xlsx")


def load_enums(workbook):
    """Return {field: set(lowercased allowed atoms)} from the Enumerations tab."""
    from openpyxl import load_workbook
    wb = load_workbook(workbook, data_only=True, read_only=True)
    ws = wb["Enumerations"]
    enums = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        field, value = row[0], row[1]
        if field == "Field" and value == "Allowed Value":
            header_seen = True
            continue
        if not header_seen or value is None:
            continue
        atoms = [a.strip().lower() for a in str(value).split("/")]  # split compound entries
        enums.setdefault(str(field), set()).update(a for a in atoms if a)
    return enums


def validate(csv_path, workbook, check=False):
    enums = load_enums(workbook)
    with open(csv_path) as f:
        rows = list(csv.DictReader(f))
    cols = rows[0].keys() if rows else []

    errors, warnings = [], []
    for i, row in enumerate(rows, start=2):  # row 1 = header
        for field in HARD_FIELDS:
            if field not in cols:
                continue
            raw = (row[field] or "").strip()
            if raw.lower() in SENTINELS:
                continue
            allowed = enums.get(field)
            if allowed is None:
                continue  # no enum defined for this field -> not controlled
            if raw.lower() not in allowed:
                near = difflib.get_close_matches(raw.lower(), allowed, n=1)
                errors.append((i, field, raw, near[0] if near else "?"))
        for field in SOFT_FIELDS:
            if field not in cols:
                continue
            raw = (row[field] or "").strip()
            if raw.lower() in SENTINELS:
                continue
            allowed = enums.get(field, set())
            if raw.lower() not in allowed:
                warnings.append((i, field, raw))

    fields_checked = [f for f in HARD_FIELDS if f in cols and f in enums]
    print(f"validated {len(rows)} rows | {os.path.basename(csv_path)}")
    print(f"hard enum fields checked: {', '.join(fields_checked)}")

    if warnings:
        seen = {}
        for _, field, raw in warnings:
            seen.setdefault((field, raw), 0)
            seen[(field, raw)] += 1
        print(f"\nSOFT warnings (PricingUnit not in canonical set - native unit passthrough):")
        for (field, raw), n in sorted(seen.items()):
            print(f"  {field}='{raw}'  x{n}")

    if errors:
        print(f"\nHARD violations: {len(errors)}")
        for r, field, raw, suggest in errors:
            print(f"  row {r}: {field}='{raw}'  ->  did you mean '{suggest}'?")
        print("\nFAIL: output is not schema-valid.")
        if check:
            sys.exit(1)
    else:
        print("\nPASS: all hard enum fields carry canonical values.")


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    csv_path = args[0] if args else "normalized_output.csv"
    wb = DEFAULT_WB
    if "--workbook" in sys.argv:
        wb = sys.argv[sys.argv.index("--workbook") + 1]
    validate(csv_path, wb, check="--check" in sys.argv)
